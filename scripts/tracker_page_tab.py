#!/usr/bin/env python3
"""
tracker_page_tab.py — per-page tabs in EA-CONTENT-TRACKER.xlsx.

team_00 rule (2026-08-17): every page IN WORK gets its own tab listing that
page's individual items — the defect, the required content, the fix, and who
must decide when it is not clear. When the page is approved the tab is HIDDEN,
never deleted: the record of how the page got approved has to survive.

The item grid encodes the work-layer rule as data:
  סיווג = «ברור»     → defect + required content + fix are all clear → we execute
  סיווג = «לא ברור»  → escalate; «הכרעה נדרשת מ» names נימרוד or אייל, and
                       «אפשרויות לבחירה» carries the options they choose between.

Usage (repo root):
    python3 scripts/tracker_page_tab.py --create R1-01
    python3 scripts/tracker_page_tab.py --create R1-01 --items items.json
    python3 scripts/tracker_page_tab.py --hide R1-01
    python3 scripts/tracker_page_tab.py --list
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_render as R  # noqa: E402
import tracker_schema as S  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
TRACKER = REPO / S.TRACKER_DIR / S.TRACKER_FILENAME

def norm(v) -> str:
    return '' if v is None else str(v).strip()


def find_page_row(wb, key: str):
    for sheet in S.DATA_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next((r for r in range(1, min(ws.max_row, 12) + 1)
                    if norm(ws.cell(r, 1).value) == S.COL_KEY),
                   S.ROUND_HEADER_ROW)
        for r in range(hdr + 1, ws.max_row + 1):
            if norm(ws.cell(r, 1).value) == key:
                return sheet, r, ws
    raise SystemExit(f'שורת עמוד «{key}» לא נמצאה')


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--create', metavar='PAGE_KEY')
    ap.add_argument('--items', help='JSON list of item dicts keyed by ITEM_HEADERS')
    ap.add_argument('--hide', metavar='PAGE_KEY')
    ap.add_argument('--list', action='store_true')
    ap.add_argument('--update', nargs=2, metavar=('PAGE_KEY', 'ITEM_KEY'),
                    help='update agent-owned cells on one item row')
    ap.add_argument('--set', action='append', default=[], metavar='COL=VALUE')
    ap.add_argument('--actor', default='team_100')
    args = ap.parse_args()

    if not TRACKER.exists():
        print(f'הטרקר לא נמצא: {TRACKER}', file=sys.stderr)
        return 2

    wb = load_workbook(TRACKER)

    if args.list:
        for name in wb.sheetnames:
            if name.startswith(S.PAGE_TAB_PREFIX):
                print(f'  {name}   [{wb[name].sheet_state}]')
        return 0

    if args.hide:
        sheet, r, ws = find_page_row(wb, args.hide)
        path = norm(ws.cell(r, S.HEADERS.index('נתיב') + 1).value)
        title = norm(ws.cell(r, S.HEADERS.index('כותרת') + 1).value)
        approval = norm(ws.cell(r, S.HEADERS.index(S.COL_APPROVAL_STATUS) + 1).value)
        if approval != S.AP_EYAL:
            print(f'סירוב: «{args.hide}» אינו «{S.AP_EYAL}» (כרגע: {approval!r}). '
                  'טאב עמוד מוסתר רק אחרי אישור אייל.', file=sys.stderr)
            return 1
        name = S.page_tab_name(path, title)
        if name not in wb.sheetnames:
            print(f'אין טאב לעמוד «{args.hide}»', file=sys.stderr)
            return 1
        wb[name].sheet_state = 'hidden'
        wb.save(TRACKER)
        print(f'  הוסתר: {name} (לא נמחק — הרשומה נשמרת)')
        return 0

    if args.update:
        page_key, item_key = args.update
        sheet, r, ws = find_page_row(wb, page_key)
        path = norm(ws.cell(r, S.HEADERS.index('נתיב') + 1).value)
        title = norm(ws.cell(r, S.HEADERS.index('כותרת') + 1).value)
        name = S.page_tab_name(path, title)
        if name not in wb.sheetnames:
            print(f'אין טאב לעמוד «{page_key}»', file=sys.stderr)
            return 1
        tab = wb[name]

        updates = {}
        for pair in args.set:
            if '=' not in pair:
                print(f'--set חייב להיות COL=VALUE: {pair!r}', file=sys.stderr)
                return 2
            col, val = pair.split('=', 1)
            col = col.strip()
            if col not in S.ITEM_HEADERS:
                print(f'עמודת סעיף לא מוכרת: {col!r}', file=sys.stderr)
                return 2
            if S.ITEM_OWNER_OF[col] == S.HUMAN:
                print(f'סירוב: «{col}» היא עמודה בבעלות אנוש.', file=sys.stderr)
                return 1
            updates[col] = val.strip()

        st = updates.get(S.COL_ITEM_STATUS)
        if st and st not in S.ITEM_STATUSES:
            print(f'סטטוס סעיף לא חוקי: {st!r}\nמותר: {list(S.ITEM_STATUSES)}',
                  file=sys.stderr)
            return 1

        target = None
        for rr in range(S.PAGE_FIRST_DATA_ROW, tab.max_row + 1):
            if norm(tab.cell(rr, 1).value) == item_key:
                target = rr
                break
        if not target:
            print(f'סעיף «{item_key}» לא נמצא בטאב «{name}»', file=sys.stderr)
            return 1

        changes = []
        for col, val in updates.items():
            c = tab.cell(target, S.ITEM_HEADERS.index(col) + 1)
            before = norm(c.value)
            if before != val:
                c.value = val
                changes.append(f'{col}: {before!r} ← {val!r}')
        if not changes:
            print('  אין שינוי.')
            return 0

        log = wb[S.SHEET_LOG]
        lr = log.max_row + 1
        for col, val in enumerate((dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                   args.actor, 'עדכון סעיף',
                                   f'{name}!{item_key}', ' | '.join(changes)), start=1):
            log.cell(lr, col, val)
        wb.save(TRACKER)
        print(f'  {name}!{item_key} — {len(changes)} שינויים')
        for ch in changes:
            print(f'    {ch}')
        return 0

    if not args.create:
        ap.print_help()
        return 2

    sheet, r, ws = find_page_row(wb, args.create)
    path = norm(ws.cell(r, S.HEADERS.index('נתיב') + 1).value)
    title = norm(ws.cell(r, S.HEADERS.index('כותרת') + 1).value)
    name = S.page_tab_name(path, title)

    items = []
    if args.items:
        items = json.loads(Path(args.items).read_text(encoding='utf-8'))
        for it in items:
            bad = [k for k in it if k not in S.ITEM_HEADERS]
            if bad:
                print(f'עמודות לא מוכרות בפריט: {bad}', file=sys.stderr)
                return 2
            for k in it:
                if S.ITEM_OWNER_OF[k] == S.HUMAN and norm(it[k]):
                    print(f'סירוב: הפריט מנסה לאכלס «{k}» — עמודה בבעלות אנוש.',
                          file=sys.stderr)
                    return 1

    if name in wb.sheetnames:
        print(f'סירוב: הטאב «{name}» כבר קיים. '
              'עדכון סעיפים קיימים נעשה בעריכה, לא בבנייה מחדש.', file=sys.stderr)
        return 1

    R.write_page_tab(wb.create_sheet(name), args.create, path, title, items)

    log = wb[S.SHEET_LOG]
    lr = log.max_row + 1
    for col, val in enumerate((dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                               args.actor, 'יצירת טאב עמוד', args.create,
                               f'{name} · {len(items)} סעיפים'), start=1):
        log.cell(lr, col, val)

    wb.save(TRACKER)
    waiting = sum(1 for i in items
                  if i.get(S.COL_ITEM_STATUS) in S.ITEM_STATUS_REQUIRING_DECIDER)
    print(f'  נוצר טאב «{name}» · {len(items)} סעיפים · {waiting} ממתינים להכרעה')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
