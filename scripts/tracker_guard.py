#!/usr/bin/env python3
"""
tracker_guard.py — the real lock on EA-CONTENT-TRACKER.xlsx.

Worksheet protection and dropdowns steer a human in the UI. They do nothing to
an agent, because openpyxl ignores protection. This script is what actually
enforces the two-permission split: it diffs the workbook against the last
committed CSV snapshot and rejects illegal edits from either side.

Modes
  ingest  (session start) — human edits are expected; they are reported as the
          work list. Hand edits to agent columns are reported as warnings.
  verify  (before commit) — the agent must not have touched a human column and
          must not have set a human-only status. Any such change FAILS.

Both modes always enforce the structural invariants: sheet set, exact header
row, unique and immutable «#» keys, no row deletion, legal status vocabulary,
legal machine-status transitions, and a written reason for «הוקפא».

Usage (repo root):
    python3 scripts/tracker_guard.py --mode ingest
    python3 scripts/tracker_guard.py --mode verify
Exit codes: 0 = clean · 1 = violation (artifact written) · 2 = cannot run
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_schema as S  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
TRACKER = REPO / S.TRACKER_DIR / S.TRACKER_FILENAME
SNAPDIR = REPO / S.SNAPSHOT_DIR
BASELINE = SNAPDIR / 'latest.csv'


def norm(v) -> str:
    """Empty cell, None and whitespace all mean «not filled in»."""
    return '' if v is None else str(v).strip()


def read_workbook() -> dict[tuple[str, str], dict[str, str]]:
    """-> {(sheet, key): {header: value}}. Raises on structural breakage."""
    wb = load_workbook(TRACKER, data_only=True)

    missing = [s for s in (S.SHEET_README, S.SHEET_STATE, S.SHEET_ROUND1,
                           S.SHEET_ROUND2, S.SHEET_ROUND3, S.SHEET_LOG)
                if s not in wb.sheetnames]
    if missing:
        raise ValueError(f'טאבים חסרים: {missing}')

    out: dict[tuple[str, str], dict[str, str]] = {}
    for sheet in S.DATA_SHEETS:
        if sheet not in wb.sheetnames:
            raise ValueError(f'טאב חסר: {sheet}')
        ws = wb[sheet]
        # Sheets carry a definition block above the grid, so locate the header
        # row rather than assuming it. Scanning also means a layout change does
        # not silently turn into a false "headers were altered" rejection.
        hdr = None
        for r in range(1, min(ws.max_row, 12) + 1):
            if norm(ws.cell(r, 1).value) == S.COL_KEY:
                hdr = r
                break
        if hdr is None:
            raise ValueError(f'לא נמצאה שורת כותרות בטאב «{sheet}»')
        headers = tuple(norm(ws.cell(hdr, c).value)
                        for c in range(1, len(S.HEADERS) + 1))
        if headers != S.HEADERS:
            raise ValueError(
                f'שורת הכותרות בטאב «{sheet}» שונתה.\n'
                f'  צפוי:  {list(S.HEADERS)}\n'
                f'  נמצא:  {list(headers)}')
        seen: set[str] = set()
        for r in range(hdr + 1, ws.max_row + 1):
            key = norm(ws.cell(r, 1).value)
            if not key:
                continue
            if key in seen:
                raise ValueError(f'מפתח «#» כפול בטאב «{sheet}»: {key}')
            seen.add(key)
            out[(sheet, key)] = {
                h: norm(ws.cell(r, c).value)
                for c, h in enumerate(S.HEADERS, start=1)
            }
    return out


def read_baseline() -> dict[tuple[str, str], dict[str, str]] | None:
    if not BASELINE.exists():
        return None
    out: dict[tuple[str, str], dict[str, str]] = {}
    with BASELINE.open(encoding='utf-8-sig', newline='') as fh:
        for row in csv.DictReader(fh):
            sheet = row.pop('__sheet__')
            out[(sheet, row[S.COL_KEY])] = {h: norm(row.get(h)) for h in S.HEADERS}
    return out


def check_conflicted_copies() -> list[str]:
    """Drive names a losing side «... (1)» / «...conflict...» — never merge silently."""
    folder = TRACKER.parent
    stem = TRACKER.stem
    hits = []
    for p in folder.glob(f'{stem}*'):
        if p.name == TRACKER.name:
            continue
        if p.suffix.lower() in ('.xlsx', '.xls'):
            hits.append(p.name)
    return hits


def diff(cur: dict, base: dict) -> tuple[list, list, list]:
    """-> (human_changes, agent_changes, deleted_keys)"""
    human, agent = [], []
    for k, row in cur.items():
        if k not in base:
            continue
        for h in S.HEADERS:
            before, after = base[k].get(h, ''), row.get(h, '')
            if before != after:
                rec = (k[0], k[1], h, before, after)
                (human if S.OWNER_OF[h] == S.HUMAN else agent).append(rec)
    deleted = [k for k in base if k not in cur]
    return human, agent, deleted


def intrinsic_violations(cur: dict, base: dict | None) -> list[str]:
    """Rules that hold regardless of who edited."""
    bad = []
    for (sheet, key), row in sorted(cur.items()):
        where = f'{sheet}!{key}'

        ms = row[S.COL_MACHINE_STATUS]
        if ms and ms not in S.MACHINE_STATUSES:
            bad.append(f'{where}: «סטטוס מכונה» ערך לא חוקי — {ms!r}')

        ap = row[S.COL_APPROVAL_STATUS]
        if ap and ap not in S.APPROVAL_STATUSES:
            bad.append(f'{where}: «סטטוס אישור» ערך לא חוקי — {ap!r}')

        if ms in S.STATUS_REQUIRING_REASON and not row[S.COL_AGENT_NOTES]:
            bad.append(f'{where}: סטטוס «{ms}» ללא סיבה כתובה בעמודת «{S.COL_AGENT_NOTES}»')

        if base and (sheet, key) in base:
            prev = base[(sheet, key)][S.COL_MACHINE_STATUS]
            if prev and ms and ms != prev:
                allowed = S.MACHINE_TRANSITIONS.get(prev, set())
                if ms not in allowed:
                    bad.append(
                        f'{where}: מעבר סטטוס אסור «{prev}» ← «{ms}». '
                        f'מותר מ«{prev}»: {sorted(allowed)}')
    return bad


def write_reject(problems: list[str], mode: str) -> Path:
    SNAPDIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime('%Y-%m-%dT%H%M%S')
    path = SNAPDIR.parent / f'TRACKER-GUARD-REJECT-{stamp}.md'
    path.write_text(
        f'# TRACKER-GUARD REJECT — {stamp}\n\n'
        f'**מצב:** `{mode}` · **קובץ:** `{S.TRACKER_DIR}/{S.TRACKER_FILENAME}`\n\n'
        '## הפרות\n\n'
        + '\n'.join(f'{i}. {p}' for i, p in enumerate(problems, 1))
        + '\n\n## מה עכשיו\n\n'
        'הטרקר לא עודכן ולא נעשה קומיט. יש לתקן את ההפרות למעלה, או — אם השינוי נכון '
        'ומכוון — להביא אותו לאישור team_00 לפני שממשיכים.\n',
        encoding='utf-8')
    return path


def main() -> int:
    global TRACKER, BASELINE
    ap = argparse.ArgumentParser()
    ap.add_argument('--mode', choices=('ingest', 'verify'), default='verify')
    ap.add_argument('--file', help='workbook to check (self-test / scratch copy)')
    ap.add_argument('--baseline', help='baseline CSV to diff against')
    args = ap.parse_args()

    if args.file:
        TRACKER = Path(args.file)
    if args.baseline:
        BASELINE = Path(args.baseline)

    if not TRACKER.exists():
        print(f'הטרקר לא נמצא: {TRACKER}', file=sys.stderr)
        return 2

    conflicts = check_conflicted_copies()
    if conflicts:
        p = write_reject(
            [f'נמצא עותק מתנגש בתיקייה המסונכרנת: «{c}». '
             'דרייב יצר אותו כששני צדדים ערכו במקביל. אין למזג לבד — '
             'יש להגיש את ההפרש ל-team_00 ולהכריע.' for c in conflicts], args.mode)
        print(f'FAIL — עותק מתנגש. artifact: {p}', file=sys.stderr)
        return 1

    try:
        cur = read_workbook()
    except ValueError as e:
        p = write_reject([str(e)], args.mode)
        print(f'FAIL — מבנה. artifact: {p}', file=sys.stderr)
        return 1

    base = read_baseline()
    problems = intrinsic_violations(cur, base)

    if base is None:
        print(f'  אין baseline — הרצה ראשונה. {len(cur)} שורות נקראו.')
        if problems:
            p = write_reject(problems, args.mode)
            print(f'FAIL — {len(problems)} הפרות. artifact: {p}', file=sys.stderr)
            return 1
        print('  PASS (מבנה + אוצר מילים תקינים)')
        return 0

    human, agent, deleted = diff(cur, base)

    if deleted:
        problems.append(
            'שורות נמחקו מהטרקר: ' + ', '.join(f'{s}!{k}' for s, k in deleted)
            + '. מחיקת שורה אסורה — שורה שיצאה מהיקף עוברת ל«הוקפא» עם סיבה.')

    if args.mode == 'verify' and human:
        for sheet, key, col, before, after in human:
            problems.append(
                f'{sheet}!{key}: הסוכן שינה את «{col}» — עמודה בבעלות אנוש. '
                f'{before!r} ← {after!r}')

    if problems:
        p = write_reject(problems, args.mode)
        print(f'FAIL — {len(problems)} הפרות. artifact: {p}', file=sys.stderr)
        for x in problems:
            print(f'  · {x}', file=sys.stderr)
        return 1

    if args.mode == 'ingest':
        new_keys = [k for k in cur if k not in base]
        print(f'  PASS · {len(cur)} שורות · {len(new_keys)} חדשות')
        if human:
            print(f'\n  קלט אנושי מאז התצלום האחרון — {len(human)} שינויים:')
            for sheet, key, col, before, after in human:
                print(f'    {sheet}!{key}  {col}: {before!r} ← {after!r}')
        else:
            print('  אין קלט אנושי חדש.')
        if agent:
            print(f'\n  ! {len(agent)} שינויי-יד בעמודות סוכן (לא חוסם, לידיעה):')
            for sheet, key, col, before, after in agent[:20]:
                print(f'    {sheet}!{key}  {col}: {before!r} ← {after!r}')
    else:
        print(f'  PASS · {len(cur)} שורות · {len(agent)} שינויי סוכן · '
              'אפס נגיעות בעמודות אנוש')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
