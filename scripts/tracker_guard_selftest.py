#!/usr/bin/env python3
"""Regression test for tracker_guard.py — proves the two-permission lock is real.

Each case mutates a scratch copy of the tracker and asserts the guard's verdict.
Run from the repo root:  python3 scripts/tracker_guard_selftest.py

Reject artifacts produced by the failing cases are cleaned up at the end,
so a passing run leaves the repo exactly as it found it.
"""
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / 'EyalAmit_Site_GoogleDrive_Sync/EA-CONTENT-TRACKER.xlsx'
BASE = REPO / '_COMMUNICATION/team_100/S006/tracker/latest.csv'
sys.path.insert(0, str(REPO / 'scripts'))
import tracker_schema as S  # noqa: E402

SHEET = S.SHEET_ROUND1
# Derived from the schema, never hardcoded: the column ORDER is deliberately
# changeable (status and responsibility lead), and a stale index would silently
# assert against the wrong cell.
COL = {h: i + 1 for i, h in enumerate(S.HEADERS)}

tmp = Path(tempfile.mkdtemp())
results = []


def header_row(ws) -> int:
    for r in range(1, 12):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() == S.COL_KEY:
            return r
    return 1


def clean_row() -> int:
    """A row still at «טרם נבדק» with empty agent notes.

    Cases that assert on status transitions must not hardcode a row number:
    as the milestone progresses, real rows advance and what was an illegal
    transition from «טרם נבדק» becomes legal from «בעבודה». Locating an
    untouched row keeps these assertions meaningful for the whole milestone.
    """
    ws = load_workbook(SRC, data_only=True)[SHEET]
    for r in range(header_row(ws) + 1, ws.max_row + 1):
        status = (ws.cell(r, COL['סטטוס מכונה']).value or '').strip()
        notes = (ws.cell(r, COL['הערות סוכן']).value or '').strip()
        if status == 'טרם נבדק' and not notes:
            return r
    raise SystemExit('no untouched row left — self-test needs one to assert on')


ROW = clean_row()


def run(path, mode='verify'):
    p = subprocess.run(
        [sys.executable, str(REPO / 'scripts/tracker_guard.py'),
         '--mode', mode, '--file', str(path), '--baseline', str(BASE)],
        capture_output=True, text=True, cwd=REPO)
    return p.returncode, (p.stdout + p.stderr)


def case(name, mutate, expect_rc, mode='verify', expect_text=None):
    path = tmp / f'{name}.xlsx'
    shutil.copy(SRC, path)
    wb = load_workbook(path)
    mutate(wb)
    wb.save(path)
    rc, out = run(path, mode)
    ok = (rc == expect_rc) and (expect_text is None or expect_text in out)
    results.append((name, ok, rc, expect_rc, out.strip().splitlines()[:3]))
    return ok


# a) agent writes a human-only approval status
case('a_agent_sets_eyal_approval',
     lambda wb: wb[SHEET].cell(ROW, COL['סטטוס אישור'], 'אושר ע״י אייל'),
     1, expect_text='עמודה בבעלות אנוש')

# b) agent writes into a human notes column
case('b_agent_writes_nimrod_notes',
     lambda wb: wb[SHEET].cell(ROW, COL['הערות נימרוד'], 'טקסט שהסוכן המציא'),
     1, expect_text='עמודה בבעלות אנוש')

# c) illegal machine transition: טרם נבדק -> הוגש לבדיקה (skips בעבודה)
case('c_illegal_transition',
     lambda wb: wb[SHEET].cell(ROW, COL['סטטוס מכונה'], 'הוגש לבדיקה'),
     1, expect_text='מעבר סטטוס אסור')

# d) frozen without a written reason
case('d_frozen_without_reason',
     lambda wb: wb[SHEET].cell(ROW, COL['סטטוס מכונה'], 'הוקפא'),
     1, expect_text='ללא סיבה כתובה')

# e) row deletion
case('e_row_deleted', lambda wb: wb[SHEET].delete_rows(ROW), 1,
     expect_text='שורות נמחקו')

# f) header renamed — must target the real header row, which sits below the
# round's definition block, not row 1.
def rename_header(wb):
    ws = wb[SHEET]
    ws.cell(header_row(ws), COL['סטטוס אישור'], 'סטטוס')


case('f_header_renamed', rename_header, 1, expect_text='FAIL — מבנה')


# f2) the key header itself renamed — guard must not silently skip the sheet
def rename_key(wb):
    ws = wb[SHEET]
    ws.cell(header_row(ws), 1, 'מזהה')


case('f2_key_header_renamed', rename_key, 1, expect_text='FAIL — מבנה')

# g) invalid status value entirely
case('g_bogus_status',
     lambda wb: wb[SHEET].cell(ROW, COL['סטטוס מכונה'], 'בערך מוכן'),
     1, expect_text='ערך לא חוקי')


# h) LEGAL agent flow: טרם נבדק -> בעבודה
def legal(wb):
    wb[SHEET].cell(ROW, COL['סטטוס מכונה'], 'בעבודה')
    wb[SHEET].cell(ROW, COL['הערות סוכן'], 'התחלנו מול סקירה דף הבית.xlsx')


case('h_legal_agent_progress', legal, 0)


# i) LEGAL frozen WITH a reason
def frozen_ok(wb):
    wb[SHEET].cell(ROW, COL['סטטוס מכונה'], 'הוקפא')
    wb[SHEET].cell(ROW, COL['הערות סוכן'], 'חסר חומר מאייל לפרק הווידאו')


case('i_legal_frozen_with_reason', frozen_ok, 0)

# j) human edit under ingest mode is accepted and surfaced as work
case('j_human_edit_ingest',
     lambda wb: wb[SHEET].cell(ROW, COL['סטטוס אישור'], 'חזר לתיקונים'),
     0, mode='ingest', expect_text='קלט אנושי')

# k) same human edit under verify mode is a violation
case('k_human_edit_verify',
     lambda wb: wb[SHEET].cell(ROW, COL['סטטוס אישור'], 'חזר לתיקונים'),
     1, mode='verify')

print(f'\nclean row used for transition assertions: {ROW}')
print('=' * 74)
passed = sum(1 for _, ok, *_ in results if ok)
for name, ok, rc, exp, head in results:
    print(f'{"PASS" if ok else "FAIL"}  {name:34} rc={rc} (expected {exp})')
    if not ok:
        for line in head:
            print(f'        {line}')
print('=' * 74)
print(f'{passed}/{len(results)} cases passed')
shutil.rmtree(tmp, ignore_errors=True)
for stray in (REPO / '_COMMUNICATION/team_100/S006').glob('TRACKER-GUARD-REJECT-*.md'):
    stray.unlink()
sys.exit(0 if passed == len(results) else 1)
