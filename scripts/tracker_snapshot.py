#!/usr/bin/env python3
"""
tracker_snapshot.py — write the git-tracked audit trail of EA-CONTENT-TRACKER.xlsx.

The xlsx in the Drive-synced folder is the SSOT and is gitignored. This script
mirrors it to CSV inside the repo so that:
  · every session leaves a diffable record of what changed and who changed it,
  · tracker_guard.py has a baseline to compare the next edit against.

The CSV is never written back into the xlsx. It is a photograph, not a source.

Usage (repo root):
    python3 scripts/tracker_snapshot.py                 # dated snapshot + latest.csv
    python3 scripts/tracker_snapshot.py --baseline-only # refresh latest.csv only
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_schema as S  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
TRACKER = REPO / S.TRACKER_DIR / S.TRACKER_FILENAME
SNAPDIR = REPO / S.SNAPSHOT_DIR

FIELDS = ('__sheet__',) + S.HEADERS
ITEM_FIELDS = ('__sheet__', '__page__') + S.ITEM_HEADERS


def norm(v) -> str:
    return '' if v is None else str(v).strip()


def rows() -> list[dict[str, str]]:
    wb = load_workbook(TRACKER, data_only=True)
    out = []
    for sheet in S.DATA_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = next((r for r in range(1, min(ws.max_row, 12) + 1)
                    if norm(ws.cell(r, 1).value) == S.COL_KEY), 1)
        for r in range(hdr + 1, ws.max_row + 1):
            if not norm(ws.cell(r, 1).value):
                continue
            rec = {'__sheet__': sheet}
            rec.update({h: norm(ws.cell(r, c).value)
                        for c, h in enumerate(S.HEADERS, start=1)})
            out.append(rec)
    return out


def item_rows() -> list[dict[str, str]]:
    """Per-page item grids. These carry the actual decisions — what was found,
    what was decided, and by whom — so they belong in the audit trail just as
    much as the page-level index does."""
    import re
    wb = load_workbook(TRACKER, data_only=True)
    out = []
    for name in wb.sheetnames:
        if not name.startswith(S.PAGE_TAB_PREFIX):
            continue
        ws = wb[name]
        m = re.search(r'שורת אב (\S+)', norm(ws.cell(1, 1).value))
        page = m.group(1) if m else name
        for r in range(S.PAGE_FIRST_DATA_ROW, ws.max_row + 1):
            if not norm(ws.cell(r, 1).value):
                continue
            rec = {'__sheet__': name, '__page__': page}
            rec.update({h: norm(ws.cell(r, c).value)
                        for c, h in enumerate(S.ITEM_HEADERS, start=1)})
            out.append(rec)
    return out


def write_csv(path: Path, data: list[dict[str, str]], fields=FIELDS) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(data)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--baseline-only', action='store_true')
    args = ap.parse_args()

    if not TRACKER.exists():
        print(f'הטרקר לא נמצא: {TRACKER}', file=sys.stderr)
        return 2

    data = rows()
    write_csv(SNAPDIR / 'latest.csv', data)
    print(f'  baseline: {SNAPDIR.relative_to(REPO)}/latest.csv ({len(data)} שורות)')

    items = item_rows()
    write_csv(SNAPDIR / 'latest-items.csv', items, ITEM_FIELDS)
    print(f'  items:    {SNAPDIR.relative_to(REPO)}/latest-items.csv ({len(items)} סעיפים)')

    if not args.baseline_only:
        stamp = dt.date.today().isoformat()
        write_csv(SNAPDIR / f'EA-CONTENT-TRACKER-{stamp}.csv', data)
        write_csv(SNAPDIR / f'EA-CONTENT-ITEMS-{stamp}.csv', items, ITEM_FIELDS)
        print(f'  snapshot: EA-CONTENT-TRACKER-{stamp}.csv + EA-CONTENT-ITEMS-{stamp}.csv')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
