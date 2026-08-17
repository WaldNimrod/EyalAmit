#!/usr/bin/env python3
"""
tracker_render.py — shared rendering for EA-CONTENT-TRACKER.xlsx.

One place that knows how a sheet is drawn, so tracker_rebuild.py and
tracker_page_tab.py cannot drift into different layouts.

Layout rules (team_00 2026-08-17):
  · status and responsibility columns come first
  · every status-like column is a dropdown
  · NO legend row inside the grid — it would be dragged around by a header
    sort and would sit inside the autofilter range. Ownership is carried by
    header colour and spelled out in README.
  · autofilter covers every column, so any header sorts and filters
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_schema as S  # noqa: E402

FILL_AGENT_HDR = PatternFill('solid', fgColor='2F4858')   # deep slate
FILL_HUMAN_HDR = PatternFill('solid', fgColor='A15C00')   # amber — «yours»
FILL_AGENT = PatternFill('solid', fgColor='E8EEF4')
FILL_HUMAN = PatternFill('solid', fgColor='FDF3E3')
FILL_WAIT = PatternFill('solid', fgColor='FBE3E0')
THIN = Side(style='thin', color='B7C4CF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROUND_WIDTHS = {
    '#': 8, 'סטטוס מכונה': 15, 'סטטוס אישור': 16, 'ממתין ל': 13,
    'נתיב': 30, 'כותרת': 32, 'סוג': 11, 'קובץ תוכן': 26,
    'מקור חומר (אייל)': 26, 'WP קשור': 22, 'תאריך עבודה אחרון': 15,
    'ראיות QA': 46, 'הערות סוכן': 46, 'הערות נימרוד': 30,
    'הערות אייל': 30, 'תאריך אישור': 14,
}
ITEM_WIDTHS = {
    '#': 7, 'מה נדרש ממך': 62, 'בחירה': 30, 'מילוי': 46, 'הערות אייל': 30,
    'סטטוס סעיף': 18, 'הכרעה נדרשת מ': 14, 'סיווג': 10,
    'הסעיף': 28, 'סקשן אצל אייל': 13, 'הכשל': 46, 'התוכן הדרוש': 46,
    'התיקון': 38, 'נתיב קוד': 32, 'קישור': 40, 'אפשרויות לבחירה': 44,
    'הערות סוכן': 32, 'הערות נימרוד': 28, 'תאריך הכרעה': 13,
}
WRAP = ('ראיות QA', 'הערות סוכן', 'הערות נימרוד', 'הערות אייל', 'כותרת',
        'הכשל', 'התוכן הדרוש', 'התיקון', 'אפשרויות לבחירה', 'הסעיף', 'נתיב קוד',
        'מה נדרש ממך', 'מילוי', 'קישור')

# column -> allowed values. Everything status-like is a dropdown.
ROUND_DROPDOWNS = {
    S.COL_MACHINE_STATUS: S.MACHINE_STATUSES,
    S.COL_APPROVAL_STATUS: S.APPROVAL_STATUSES,
    S.COL_WAITING_ON: S.WAITING_ON,
    'סוג': S.ROW_TYPES,
}
ITEM_DROPDOWNS = {
    S.COL_ITEM_STATUS: S.ITEM_STATUSES,
    S.COL_ITEM_CLASS: S.ITEM_CLASSES,
    S.COL_ITEM_DECIDER: S.DECIDERS,
}


def write_header(ws, headers, owner_of, widths, row: int) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row, col, h)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = FILL_AGENT_HDR if owner_of[h] == S.AGENT else FILL_HUMAN_HDR
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 18)
    ws.row_dimensions[row].height = 30


def add_dropdowns(ws, headers, mapping, first: int, last: int) -> None:
    for colname, values in mapping.items():
        if colname not in headers:
            continue
        dv = DataValidation(type='list',
                            formula1='"' + ','.join(values) + '"',
                            allow_blank=True, showDropDown=False,
                            errorTitle='ערך לא חוקי',
                            error=f'«{colname}» מוגבלת לרשימה הסגורה.')
        ws.add_data_validation(dv)
        L = get_column_letter(headers.index(colname) + 1)
        dv.add(f'{L}{first}:{L}{max(last, first)}')


def finish(ws, headers, hdr_row: int, first: int, last: int) -> None:
    """Freeze below the header and make every column sortable/filterable."""
    ws.freeze_panes = ws.cell(first, 1).coordinate
    ws.auto_filter.ref = (f'A{hdr_row}:'
                          f'{get_column_letter(len(headers))}{max(last, first)}')
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.sort = False        # sorting must stay available
    ws.protection.autoFilter = False  # filtering must stay available


def write_round_sheet(ws, sheet_name: str, rows: list[dict]) -> None:
    ws.sheet_view.rightToLeft = True
    d = S.ROUND_DEFINITIONS.get(sheet_name)

    # Definition block sits ABOVE the grid in its own rows, never inside it.
    if d:
        ws.cell(1, 1, d['title']).font = Font(bold=True, size=14, color='2F4858')
        ws.cell(1, 3, d['opens']).font = Font(bold=True, size=10, color='A15C00')
        ws.cell(2, 1, d['scope']).font = Font(size=10, color='44586B')
        ws.cell(2, 8, 'תנאי סיום: ' + d['exit']).font = Font(size=9, italic=True,
                                                             color='6B7C8C')
    hdr = 3 if d else S.ROUND_HEADER_ROW
    first = hdr + 1

    write_header(ws, list(S.HEADERS), S.OWNER_OF, ROUND_WIDTHS, hdr)
    for i, values in enumerate(rows):
        r = first + i
        waiting = values.get(S.COL_WAITING_ON) in (S.W_NIMROD, S.W_EYAL)
        for col, h in enumerate(S.HEADERS, start=1):
            val = values.get(h, '')
            if h in S.SUMMARY_COLUMNS:
                val = S.summarize(val)
            c = ws.cell(r, col, val)
            if S.OWNER_OF[h] == S.HUMAN:
                c.fill = FILL_HUMAN
            else:
                c.fill = FILL_WAIT if waiting else FILL_AGENT
            c.border = BORDER
            c.alignment = Alignment(horizontal='right', vertical='top',
                                    wrap_text=h in WRAP)
            c.protection = Protection(locked=(S.OWNER_OF[h] == S.AGENT))
    last = first + max(len(rows), 1) - 1
    add_dropdowns(ws, list(S.HEADERS), ROUND_DROPDOWNS, first, last)
    finish(ws, list(S.HEADERS), hdr, first, last)


def write_page_tab(ws, page_key: str, path: str, title: str,
                   items: list[dict]) -> None:
    ws.sheet_view.rightToLeft = True
    ws.cell(1, 1, f'{title}   ·   {path}   ·   שורת אב {page_key}')
    ws.cell(1, 1).font = Font(bold=True, size=13, color='2F4858')
    ws.cell(2, 1, 'סיווג «ברור» = מבוצע באורקסטרציה ללא שאלה.  '
                  '«לא ברור» = אסקלציה — «הכרעה נדרשת מ» + «אפשרויות לבחירה».  '
                  'עמודות בכותרת כתומה הן שלכם בלבד.')
    ws.cell(2, 1).font = Font(size=9, italic=True, color='6B7C8C')

    hdr = S.PAGE_HEADER_ROW
    first = S.PAGE_FIRST_DATA_ROW
    write_header(ws, list(S.ITEM_HEADERS), S.ITEM_OWNER_OF, ITEM_WIDTHS, hdr)
    for i, item in enumerate(items):
        r = first + i
        waiting = item.get(S.COL_ITEM_STATUS) in S.ITEM_STATUS_REQUIRING_DECIDER
        for col, h in enumerate(S.ITEM_HEADERS, start=1):
            c = ws.cell(r, col, item.get(h, ''))  # '_picks' is not a header, never written
            if S.ITEM_OWNER_OF[h] == S.HUMAN:
                c.fill = FILL_HUMAN
            else:
                c.fill = FILL_WAIT if waiting else FILL_AGENT
            c.border = BORDER
            c.alignment = Alignment(horizontal='right', vertical='top',
                                    wrap_text=h in WRAP)
            c.protection = Protection(locked=(S.ITEM_OWNER_OF[h] == S.AGENT))
    last = first + max(len(items), 1) - 1
    add_dropdowns(ws, list(S.ITEM_HEADERS), ITEM_DROPDOWNS, first, last)

    # «בחירה» gets a dropdown per ROW, built from that item's own short labels —
    # team_00: «הוא בוחר ומסמן, לא כותב». Options arrive as item['_picks'].
    pick_col = get_column_letter(S.ITEM_HEADERS.index(S.COL_ITEM_PICK) + 1)
    for i, item in enumerate(items):
        picks = item.get('_picks') or []
        if not picks:
            continue
        dv = DataValidation(type='list', formula1='"' + ','.join(picks) + '"',
                            allow_blank=True, showDropDown=False,
                            promptTitle='בחירה', prompt='בחרו מהרשימה — אין צורך לכתוב.')
        ws.add_data_validation(dv)
        dv.add(f'{pick_col}{first + i}')

    finish(ws, list(S.ITEM_HEADERS), hdr, first, last)
