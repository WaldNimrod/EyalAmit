#!/usr/bin/env python3
"""
tracker_rebuild.py — rebuild EA-CONTENT-TRACKER.xlsx onto the current schema.

Used when the LAYOUT changes (column order, new columns, new round sheets).
It preserves every existing value, including the human-owned columns, and
re-derives the columns that are computed rather than typed.

Round 2 is seeded here as a COMPLETE BACKUP of the remaining live inventory —
posts, QR pages, legal, and legacy duplicates — so no page can fall through the
cracks while the round sits closed. Rows are probed for their real HTTP status
so legacy 301s are labelled from evidence rather than from a guess.

Usage (repo root):
    python3 scripts/tracker_rebuild.py --from export.json
    python3 scripts/tracker_rebuild.py --from export.json --no-probe
"""
from __future__ import annotations

import argparse
import concurrent.futures as cf
import datetime as dt
import html
import json
import re
import ssl
import sys
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_render as R  # noqa: E402
import tracker_schema as S  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
TRACKER = REPO / S.TRACKER_DIR / S.TRACKER_FILENAME
CTX = ssl._create_unverified_context()


def fetch(kind: str) -> list[dict]:
    out = []
    for page in range(1, 5):
        url = (f'{S.STAGING_BASE}/wp-json/wp/v2/{kind}?per_page=100&page={page}'
               '&status=publish&_fields=id,link,title')
        try:
            with urllib.request.urlopen(url, context=CTX, timeout=30) as r:
                batch = json.load(r)
        except Exception:
            break
        if not batch:
            break
        out += batch
        if len(batch) < 100:
            break
    return out


def probe(path: str) -> str:
    """Real HTTP status, so «legacy/301» is evidence and not assumption."""
    req = urllib.request.Request(S.STAGING_BASE + path, method='HEAD')
    try:
        op = urllib.request.build_opener(
            type('NoRedir', (urllib.request.HTTPRedirectHandler,),
                 {'redirect_request': lambda *a, **k: None})())
        with op.open(req, timeout=15) as r:
            return str(r.status)
    except urllib.error.HTTPError as e:
        return str(e.code)
    except Exception:
        return '?'


def seed_round2(round1_paths: set[str], do_probe: bool) -> list[dict]:
    pages, posts = fetch('pages'), fetch('posts')
    rows: list[dict] = []

    def path_of(x):
        return x['link'].replace(S.STAGING_BASE, '') or '/'

    cand = []
    for p in pages:
        pt = path_of(p)
        if pt in round1_paths:
            continue
        cand.append((pt, html.unescape(re.sub('<[^>]+>', '',
                                              p['title']['rendered'])).strip(),
                     S.TYPE_QR if pt.startswith('/qr/') else S.TYPE_PAGE))
    for p in posts:
        cand.append((path_of(p), html.unescape(re.sub('<[^>]+>', '',
                                                      p['title']['rendered'])).strip(),
                     S.TYPE_POST))

    statuses: dict[str, str] = {}
    if do_probe:
        with cf.ThreadPoolExecutor(max_workers=8) as ex:
            for pt, st in zip([c[0] for c in cand],
                              ex.map(probe, [c[0] for c in cand])):
                statuses[pt] = st

    cand.sort(key=lambda c: ({S.TYPE_PAGE: 0, S.TYPE_POST: 1, S.TYPE_QR: 2}[c[2]],
                             c[0]))
    for n, (pt, title, typ) in enumerate(cand, start=1):
        st = statuses.get(pt, '')
        if st.startswith('3'):
            typ = S.TYPE_LEGACY
        note = ''
        if typ == S.TYPE_LEGACY:
            note = f'מפנה (HTTP {st}) — אימות יעד ה-301 בלבד, לא עריכת תוכן.'
        elif st and st != '200':
            note = f'⚠ HTTP {st} — לבדוק לפני עבודת תוכן.'
        rows.append({
            '#': f'R2-{n:03d}',
            S.COL_MACHINE_STATUS: S.ST_NOT_CHECKED,
            S.COL_APPROVAL_STATUS: S.AP_NONE,
            S.COL_WAITING_ON: S.W_TEAM100,
            'נתיב': pt, 'כותרת': title, 'סוג': typ,
            'קובץ תוכן': '', 'מקור חומר (אייל)': '', 'WP קשור': '',
            'תאריך עבודה אחרון': '', 'ראיות QA': '', 'הערות סוכן': note,
            'הערות נימרוד': '', 'הערות אייל': '', 'תאריך אישור': '',
        })
    return rows


def build_readme(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 120
    lines = [
        ('טרקר עדכון תוכן — eyalamit.co.il', 'h1'),
        (f'אבן דרך S006 · גרסת סכימה {S.TRACKER_VERSION}', 'sub'),
        ('', ''),
        ('הקובץ הזה הוא מקור האמת היחיד לרשימת העמודים, לסטטוס ולהערות. '
         'נמצא בתיקייה המסונכרנת ועולה לדרייב מעצמו.', 'p'),
        ('', ''),
        ('חוק התוכן', 'h2'),
        ('תוכן יכול להיות אחד משלושה בלבד: (1) מה שקיים באתר כשאין הערה · '
         '(2) מה שהתקבל מאייל בקבצים · (3) מה שמתקבל מנימרוד ישירות.', 'p'),
        ('אין ניחוש. אין כתיבה עצמית. אין השלמת פערים. חסר חומר → «הוקפא» עם סיבה.', 'p'),
        ('', ''),
        ('שתי הרשאות — לפי צבע הכותרת', 'h2'),
        ('כותרת כהה = עמודה בבעלות הסוכן. אתם קוראים אותה, לא עורכים.', 'p'),
        ('כותרת כתומה = עמודה בבעלותכם. הסוכן קורא ומגיב, ולעולם לא כותב בה.', 'p'),
        ('ניסיון של סוכן לכתוב בעמודה כתומה נחסם על ידי scripts/tracker_guard.py.', 'p'),
        ('', ''),
        ('העמודה «ממתין ל» — מי צריך לגעת בעמוד הבא', 'h2'),
        ('נגזרת אוטומטית ממצב הסעיפים בטאב העמוד: אם סעיף כלשהו ממתין לאייל → אייל; '
         'אחרת אם ממתין לנימרוד → נימרוד; אחרת לפי סטטוס העמוד.', 'p'),
        ('מיינו לפיה כדי לראות מיד על מי הכל תקוע.', 'p'),
        ('', ''),
        ('סטטוסים — כולם רשימה סגורה (dropdown)', 'h2'),
        ('סטטוס מכונה (סוכן): ' + ' · '.join(S.MACHINE_STATUSES), 'p'),
        ('סטטוס אישור (אתם): ' + ' · '.join(S.APPROVAL_STATUSES), 'p'),
        ('ממתין ל: ' + ' · '.join(S.WAITING_ON), 'p'),
        ('סטטוס סעיף (בטאבי העמודים): ' + ' · '.join(S.ITEM_STATUSES), 'p'),
        ('', ''),
        ('הערות', 'h2'),
        ('בגיליונות הסבבים ההערות הן כותרת בלבד — עד שתי שורות. '
         'הפירוט המלא נמצא תמיד בטאב של אותו עמוד.', 'p'),
        ('', ''),
        ('טאב לכל עמוד בעבודה', 'h2'),
        ('כל עמוד בעבודה מקבל טאב משלו עם כל הסעיפים שלו. '
         'עמוד שאושר — הטאב מוסתר, לא נמחק.', 'p'),
        ('', ''),
        ('מיון וסינון', 'h2'),
        ('בכל הגיליונות יש סינון על כל העמודות — לחצו על החץ בכותרת כדי למיין או לסנן.', 'p'),
        ('', ''),
        ('לפני שאתם עורכים', 'h2'),
        ('הציצו בטאב «מצב». אם LOCK מלא — סוכן עובד על הקובץ כרגע, המתינו.', 'p'),
    ]
    for i, (text, kind) in enumerate(lines, start=1):
        c = ws.cell(i, 1, text)
        c.font = {'h1': Font(bold=True, size=16, color='2F4858'),
                  'sub': Font(size=10, italic=True, color='6B7C8C'),
                  'h2': Font(bold=True, size=12, color='36618E')}.get(
                      kind, Font(size=11))
        c.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)


def build_state(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 62
    ws.cell(1, 1, 'מצב הקובץ').font = Font(bold=True, size=14, color='2F4858')
    ws.cell(2, 1, 'כש-LOCK מלא — סוכן עובד על הקובץ כרגע. אל תערכו עד שיתרוקן.'
            ).font = Font(size=10, italic=True, color='6B7C8C')
    for i, (k, v) in enumerate((('LOCK', ''), ('נעול על ידי', ''), ('מאז', ''),
                                ('עדכון אחרון', dt.date.today().isoformat()),
                                ('גרסת סכימה', S.TRACKER_VERSION)), start=4):
        a = ws.cell(i, 1, k)
        a.font = Font(bold=True)
        a.fill = R.FILL_AGENT
        a.border = R.BORDER
        b = ws.cell(i, 2, v)
        b.fill = R.FILL_AGENT
        b.border = R.BORDER


def build_log(ws, entries) -> None:
    ws.sheet_view.rightToLeft = True
    headers = ('חותמת זמן', 'מבצע', 'פעולה', 'שורות מושפעות', 'פירוט')
    for col, (h, w) in enumerate(zip(headers, (20, 14, 24, 22, 80)), start=1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = R.FILL_AGENT_HDR
        c.border = R.BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    for i, row in enumerate(entries, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(i, col, val)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{max(len(entries) + 1, 2)}'


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--from', dest='src', required=True)
    ap.add_argument('--no-probe', action='store_true')
    args = ap.parse_args()

    data = json.loads(Path(args.src).read_text(encoding='utf-8'))
    r1 = data['rounds'].get(S.SHEET_ROUND1, [])

    # item statuses per page tab, so «ממתין ל» is derived from real blockers
    items_by_page: dict[str, list[str]] = {}
    for tab, payload in data['pages'].items():
        m = re.search(r'שורת אב (\S+)', payload.get('title', ''))
        if m:
            items_by_page[m.group(1)] = [
                it.get(S.COL_ITEM_STATUS, '') for it in payload['items']]

    for row in r1:
        row[S.COL_WAITING_ON] = S.derive_waiting_on(
            row.get(S.COL_MACHINE_STATUS, ''), row.get(S.COL_APPROVAL_STATUS, ''),
            items_by_page.get(row['#'], ()))

    round1_paths = {r['נתיב'] for r in r1}
    print('  seeding round 2 from live inventory'
          + ('' if args.no_probe else ' (probing HTTP status)') + ' …')
    r2 = seed_round2(round1_paths, not args.no_probe)
    print(f'  round 2: {len(r2)} rows')

    wb = Workbook()
    build_readme(wb.active)
    wb.active.title = S.SHEET_README
    build_state(wb.create_sheet(S.SHEET_STATE))
    R.write_round_sheet(wb.create_sheet(S.SHEET_ROUND1), S.SHEET_ROUND1, r1)
    R.write_round_sheet(wb.create_sheet(S.SHEET_ROUND2), S.SHEET_ROUND2, r2)
    R.write_round_sheet(wb.create_sheet(S.SHEET_ROUND3), S.SHEET_ROUND3, [])

    for tab, payload in data['pages'].items():
        title = payload.get('title', '')
        m = re.search(r'^(.*?)\s+·\s+(\S+)\s+·\s+שורת אב (\S+)', title)
        disp, path, key = (m.group(1), m.group(2), m.group(3)) if m else (tab, '', '')
        R.write_page_tab(wb.create_sheet(tab), key, path, disp, payload['items'])

    entries = list(data.get('log', []))
    entries.append([dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'team_100',
                    'בנייה מחדש של המבנה',
                    f'{len(r1)} + {len(r2)} שורות',
                    'סדר עמודות חדש (סטטוס ואחריות ראשונים) · עמודת «ממתין ל» · '
                    'סינון על כל העמודות · הערות מקוצרות · גיבוי סבב 2 · גיליון סבב 3'])
    build_log(wb.create_sheet(S.SHEET_LOG), entries)

    wb.save(TRACKER)
    print(f'  wrote {TRACKER.name}: {len(wb.sheetnames)} sheets — '
          + ', '.join(wb.sheetnames))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
