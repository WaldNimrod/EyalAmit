#!/usr/bin/env python3
"""
Build EA-CONTENT-TRACKER.xlsx — the single source of truth for the S006
content-accuracy milestone.

The workbook is written into the Drive-synced folder, so it uploads to Drive on
its own and can be opened in Google Sheets by Nimrod and Eyal. Agents edit the
same file locally through openpyxl.

Round 1 rows are seeded from tracker_schema.ROUND1_SEED, with live titles
fetched from the staging REST API so the sheet describes what is actually
published rather than what we assume is published.

Usage (repo root):
    python3 scripts/tracker_build.py              # build round 1
    python3 scripts/tracker_build.py --round2     # populate round 2 (after round 1 closes)
    python3 scripts/tracker_build.py --dry-run
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import ssl
import sys
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_schema as S  # noqa: E402

REPO = Path(__file__).resolve().parent.parent

# Visual language — deliberately plain. This is a working document, not a deck.
FILL_AGENT = PatternFill('solid', fgColor='E8EEF4')   # cool grey-blue: machine-owned
FILL_HUMAN = PatternFill('solid', fgColor='FDF3E3')   # warm sand: human-owned
FILL_TITLE = PatternFill('solid', fgColor='2F4858')
THIN = Side(style='thin', color='B7C4CF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    '#': 6, 'סוג': 11, 'נתיב': 30, 'כותרת': 34, 'קובץ תוכן': 30,
    'מקור חומר (אייל)': 30, 'WP קשור': 24, 'סטטוס מכונה': 16,
    'תאריך עבודה אחרון': 18, 'ראיות QA': 26, 'הערות סוכן': 42,
    'סטטוס אישור': 18, 'הערות נימרוד': 38, 'הערות אייל': 38, 'תאריך אישור': 15,
}


def fetch_live(kind: str) -> list[dict]:
    """Pull published pages/posts from staging. Returns [] if unreachable."""
    ctx = ssl._create_unverified_context()
    out: list[dict] = []
    for page in range(1, 5):
        url = (f'{S.STAGING_BASE}/wp-json/wp/v2/{kind}?per_page=100&page={page}'
               '&status=publish&_fields=id,slug,link,title,date,modified')
        try:
            with urllib.request.urlopen(url, context=ctx, timeout=30) as r:
                batch = json.load(r)
        except Exception:
            break
        if not batch:
            break
        out += batch
        if len(batch) < 100:
            break
    return out


def title_index(pages: list[dict]) -> dict[str, str]:
    """path -> rendered title."""
    import html
    import re
    idx = {}
    for p in pages:
        path = p['link'].replace(S.STAGING_BASE, '') or '/'
        title = html.unescape(re.sub('<[^>]+>', '', p['title']['rendered'])).strip()
        idx[path] = title
    return idx


def style_header(ws) -> None:
    for col, (header, owner) in enumerate(S.COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.fill = FILL_TITLE
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(header, 18)
    ws.row_dimensions[1].height = 34

    # Ownership legend row directly under the header, so the split is visible
    # to anyone who opens the file without reading the README tab.
    for col, (header, owner) in enumerate(S.COLUMNS, start=1):
        label = 'סוכן' if owner == S.AGENT else 'אנוש ← רק אתם'
        c = ws.cell(row=2, column=col, value=label)
        c.font = Font(bold=True, size=9,
                      color='36618E' if owner == S.AGENT else 'A15C00')
        c.fill = FILL_AGENT if owner == S.AGENT else FILL_HUMAN
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER


def add_validations(ws, last_row: int) -> None:
    """Dropdowns on both status columns — the UX half of the lock."""
    mi = S.HEADERS.index(S.COL_MACHINE_STATUS) + 1
    ai = S.HEADERS.index(S.COL_APPROVAL_STATUS) + 1

    dv_m = DataValidation(
        type='list', formula1='"' + ','.join(S.MACHINE_STATUSES) + '"',
        allow_blank=False, showDropDown=False,
        errorTitle='ערך לא חוקי',
        error='סטטוס מכונה מוגבל לרשימה. עמודה זו בבעלות הסוכן.')
    dv_a = DataValidation(
        type='list', formula1='"' + ','.join(S.APPROVAL_STATUSES) + '"',
        allow_blank=False, showDropDown=False,
        errorTitle='ערך לא חוקי',
        error='סטטוס אישור מוגבל לרשימה, ונקבע על ידי נימרוד או אייל בלבד.')
    ws.add_data_validation(dv_m)
    ws.add_data_validation(dv_a)
    if last_row >= 3:
        dv_m.add(f'{get_column_letter(mi)}3:{get_column_letter(mi)}{last_row}')
        dv_a.add(f'{get_column_letter(ai)}3:{get_column_letter(ai)}{last_row}')


def protect(ws, last_row: int) -> None:
    """Lock agent columns in the UI; leave human columns open.

    Deliberately password-less: this is a guard rail that tells a human «not
    yours», not a security boundary. The real enforcement is tracker_guard.py,
    which is why an agent (openpyxl ignores protection) is still governed.
    """
    for row in range(3, max(last_row, 3) + 1):
        for col, (header, owner) in enumerate(S.COLUMNS, start=1):
            ws.cell(row=row, column=col).protection = Protection(
                locked=(owner == S.AGENT))
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.formatCells = False


def write_row(ws, r: int, values: dict) -> None:
    for col, (header, owner) in enumerate(S.COLUMNS, start=1):
        c = ws.cell(row=r, column=col, value=values.get(header, ''))
        c.fill = FILL_AGENT if owner == S.AGENT else FILL_HUMAN
        c.border = BORDER
        c.alignment = Alignment(
            horizontal='right', vertical='top',
            wrap_text=header in ('הערות סוכן', 'הערות נימרוד', 'הערות אייל',
                                 'כותרת', 'ראיות QA'))


def build_data_sheet(ws, rows: list[dict]) -> None:
    ws.sheet_view.rightToLeft = True
    style_header(ws)
    for i, values in enumerate(rows, start=3):
        write_row(ws, i, values)
    last = 2 + len(rows)
    add_validations(ws, last)
    protect(ws, last)
    ws.freeze_panes = 'A3'
    if rows:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(S.HEADERS))}{last}'


def build_readme(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 118
    lines = [
        ('טרקר עדכון תוכן — eyalamit.co.il', 'h1'),
        (f'אבן דרך S006 · גרסת סכימה {S.TRACKER_VERSION}', 'sub'),
        ('', ''),
        ('הקובץ הזה הוא מקור האמת היחיד לרשימת העמודים, לסטטוס ולהערות.', 'p'),
        ('נמצא בתיקייה המסונכרנת ועולה לדרייב מעצמו. גם סוכנים וגם בני אדם עורכים אותו — כל צד בעמודות שלו.', 'p'),
        ('', ''),
        ('חוק התוכן', 'h2'),
        ('תוכן יכול להיות אחד משלושה בלבד: (1) מה שקיים באתר כשאין הערה · (2) מה שהתקבל מאייל בקבצים · '
         '(3) מה שמתקבל מנימרוד ישירות במהלך העבודה.', 'p'),
        ('אין ניחוש. אין כתיבה עצמית. אין השלמת פערים. חסר חומר → השורה עוברת ל«הוקפא» עם סיבה, ולא ממציאים.', 'p'),
        ('', ''),
        ('שתי הרשאות', 'h2'),
        ('עמודות בתכלת = בבעלות הסוכן. אתם קוראים אותן, לא עורכים.', 'p'),
        ('עמודות בחול = בבעלותכם. הסוכן קורא אותן ומגיב, ולעולם לא כותב בהן.', 'p'),
        ('', ''),
        ('סטטוס מכונה — נקבע על ידי הסוכן בלבד', 'h2'),
        ('לא ידוע ← טרם נבדק ← בעבודה ← הוגש לבדיקה. בכל שלב אפשר «הוקפא» עם סיבה בעמודת הערות סוכן.', 'p'),
        ('', ''),
        ('סטטוס אישור — נקבע על ידכם בלבד', 'h2'),
        ('חזר לתיקונים · אושר ע״י נימרוד · אושר ע״י אייל · הוקפא.', 'p'),
        ('«חזר לתיקונים» מחזיר את השורה לעבודה בסבב הבא. סוכן שינסה לכתוב כאן — ייעצר.', 'p'),
        ('', ''),
        ('לפני שאתם עורכים', 'h2'),
        ('הציצו בטאב «מצב». אם הוא מראה שסוכן מחזיק את הקובץ כרגע — המתינו לסיום, אחרת דרייב עלול לייצר עותק מתנגש.', 'p'),
        ('', ''),
        ('סבבים', 'h2'),
        ('סבב 1 — ליבה: עמודי התפריט הראשי וכל עמוד שאייל סיפק לו חומר.', 'p'),
        ('סבב 2 — כל שאר העמודים הזמינים לגולש שכבר קיימים. נפתח רק בסגירת סבב 1.', 'p'),
        ('סבב 3 — תוכן תומך ואופטימיזציה. אבן דרך נפרדת (S007).', 'p'),
    ]
    for i, (text, kind) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if kind == 'h1':
            c.font = Font(bold=True, size=16, color='2F4858')
        elif kind == 'sub':
            c.font = Font(size=10, italic=True, color='6B7C8C')
        elif kind == 'h2':
            c.font = Font(bold=True, size=12, color='36618E')
        else:
            c.font = Font(size=11)
        c.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)


def build_state(ws) -> None:
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 62
    rows = [
        ('LOCK', ''),
        ('נעול על ידי', ''),
        ('מאז', ''),
        ('עדכון אחרון', dt.date.today().isoformat()),
        ('גרסת סכימה', S.TRACKER_VERSION),
    ]
    ws.cell(row=1, column=1, value='מצב הקובץ').font = Font(bold=True, size=14, color='2F4858')
    ws.cell(row=2, column=1,
            value='כש-LOCK מלא — סוכן עובד על הקובץ כרגע. אל תערכו עד שיתרוקן.'
            ).font = Font(size=10, italic=True, color='6B7C8C')
    for i, (k, v) in enumerate(rows, start=4):
        a = ws.cell(row=i, column=1, value=k)
        a.font = Font(bold=True)
        a.fill = FILL_AGENT
        a.border = BORDER
        b = ws.cell(row=i, column=2, value=v)
        b.fill = FILL_AGENT
        b.border = BORDER


def build_log(ws) -> None:
    ws.sheet_view.rightToLeft = True
    headers = ('חותמת זמן', 'מבצע', 'פעולה', 'שורות מושפעות', 'פירוט')
    widths = (20, 14, 24, 18, 70)
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = FILL_TITLE
        c.border = BORDER
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'


def seed_round1() -> list[dict]:
    pages = fetch_live('pages')
    titles = title_index(pages)
    if not titles:
        print('  ! staging REST unreachable — titles left blank, paths still seeded',
              file=sys.stderr)

    rows = []
    for n, (path, content_file, material) in enumerate(S.ROUND1_SEED, start=1):
        note = ''
        if not titles:
            note = 'כותרת לא נשלפה — סטייג׳ינג לא היה זמין בזמן הבנייה'
        elif path not in titles:
            note = 'לא נמצא ב-REST של סטייג׳ינג — לאמת שהעמוד קיים ומפורסם'
        rows.append({
            '#': f'R1-{n:02d}',
            'סוג': S.TYPE_PAGE,
            'נתיב': path,
            'כותרת': titles.get(path, ''),
            'קובץ תוכן': content_file,
            'מקור חומר (אייל)': material,
            'WP קשור': S.WP_BY_PATH.get(path, ''),
            'סטטוס מכונה': S.ST_NOT_CHECKED,
            'תאריך עבודה אחרון': '',
            'ראיות QA': '',
            'הערות סוכן': note,
            'סטטוס אישור': S.AP_NONE,
            'הערות נימרוד': '',
            'הערות אייל': '',
            'תאריך אישור': '',
        })
    # The one menu entry that is an external link, not a page.
    rows.append({
        '#': f'R1-{len(rows) + 1:02d}',
        'סוג': S.TYPE_EXTERNAL,
        'נתיב': '(תפריט) קורסים',
        'כותרת': 'קורסים — סקולר / חיצוני',
        'קובץ תוכן': 'פריט תפריט בלבד',
        'מקור חומר (אייל)': '',
        'WP קשור': '',
        'סטטוס מכונה': S.ST_NOT_CHECKED,
        'תאריך עבודה אחרון': '',
        'ראיות QA': '',
        'הערות סוכן': 'מצביע כרגע ל-# בתפריט החי — נדרשת החלטה על היעד',
        'סטטוס אישור': S.AP_NONE,
        'הערות נימרוד': '',
        'הערות אייל': '',
        'תאריך אישור': '',
    })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--round2', action='store_true',
                    help='populate the round-2 tab from live inventory')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    target = REPO / S.TRACKER_DIR / S.TRACKER_FILENAME

    if args.round2:
        print('round 2 population is opened only after round 1 closes — '
              'not implemented in this build', file=sys.stderr)
        return 2

    if target.exists():
        print(f'REFUSING: {target} already exists.\n'
              'Rebuilding would discard human columns. Delete it deliberately '
              'first if you really mean to start over.', file=sys.stderr)
        return 1

    rows = seed_round1()
    print(f'  seeded {len(rows)} round-1 rows')

    wb = Workbook()
    build_readme(wb.active)
    wb.active.title = S.SHEET_README
    build_state(wb.create_sheet(S.SHEET_STATE))
    build_data_sheet(wb.create_sheet(S.SHEET_ROUND1), rows)
    build_data_sheet(wb.create_sheet(S.SHEET_ROUND2), [])
    build_log(wb.create_sheet(S.SHEET_LOG))

    if args.dry_run:
        print(f'  dry-run — would write {target}')
        return 0

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    print(f'  wrote {target} ({os.path.getsize(target)} bytes)')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
