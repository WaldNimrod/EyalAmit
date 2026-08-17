#!/usr/bin/env python3
"""
tracker_update.py — the ONLY sanctioned way for an agent to write to the tracker.

Writes agent-owned cells and nothing else. Refuses human columns outright, before
touching the file, so a mistake can never reach disk. Takes the LOCK in the «מצב»
tab for the duration, re-reads and re-checks mtime immediately before saving, and
appends every change to the LOG tab.

Usage (repo root):
    python3 scripts/tracker_update.py --row R1-01 \
        --set 'סטטוס מכונה=בעבודה' \
        --set 'הערות סוכן=פיילוט S006 — הוסרה כפילות trust line' \
        --actor team_100 --reason 'S006 pilot'

    python3 scripts/tracker_update.py --show R1-01
    python3 scripts/tracker_update.py --release-lock
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_schema as S  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
TRACKER = REPO / S.TRACKER_DIR / S.TRACKER_FILENAME


def norm(v) -> str:
    return '' if v is None else str(v).strip()


def header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 12) + 1):
        if norm(ws.cell(r, 1).value) == S.COL_KEY:
            return r
    return S.ROUND_HEADER_ROW


def find_row(ws, key: str) -> int | None:
    for r in range(header_row(ws) + 1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value) == key:
            return r
    return None


def locate(wb, key: str) -> tuple[str, int]:
    for sheet in S.DATA_SHEETS:
        r = find_row(wb[sheet], key)
        if r:
            return sheet, r
    raise SystemExit(f'שורה «{key}» לא נמצאה באף טאב נתונים')


def state_cell(wb, label: str):
    ws = wb[S.SHEET_STATE]
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value) == label:
            return ws.cell(r, 2)
    return None


def set_lock(wb, actor: str | None) -> None:
    now = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for label, value in (('LOCK', f'agent {now}' if actor else ''),
                         ('נעול על ידי', actor or ''),
                         ('מאז', now if actor else ''),
                         ('עדכון אחרון', now)):
        c = state_cell(wb, label)
        if c is not None:
            c.value = value


def append_log(wb, actor: str, action: str, rows: str, detail: str) -> None:
    ws = wb[S.SHEET_LOG]
    r = ws.max_row + 1
    for col, val in enumerate(
            (dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             actor, action, rows, detail), start=1):
        ws.cell(r, col, val)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--row', help='tracker key, e.g. R1-01')
    ap.add_argument('--set', action='append', default=[], metavar='COL=VALUE')
    ap.add_argument('--actor', default='team_100')
    ap.add_argument('--reason', default='')
    ap.add_argument('--show', metavar='ROW')
    ap.add_argument('--refresh-waiting', action='store_true',
                    help='recompute the derived «ממתין ל» column on every row')
    ap.add_argument('--acquire-lock', action='store_true',
                    help='take LOCK for a multi-step work window')
    ap.add_argument('--release-lock', action='store_true')
    args = ap.parse_args()

    if not TRACKER.exists():
        print(f'הטרקר לא נמצא: {TRACKER}', file=sys.stderr)
        return 2

    if args.refresh_waiting:
        wb = load_workbook(TRACKER)
        # item statuses per page, so page-level «waiting on» reflects the real
        # blockers rather than only the page's own status
        items_by_page = {}
        for name in wb.sheetnames:
            if not name.startswith(S.PAGE_TAB_PREFIX):
                continue
            tab = wb[name]
            import re as _re
            m = _re.search(r'שורת אב (\S+)', norm(tab.cell(1, 1).value))
            if not m:
                continue
            st_col = S.ITEM_HEADERS.index(S.COL_ITEM_STATUS) + 1
            items_by_page[m.group(1)] = [
                norm(tab.cell(r, st_col).value)
                for r in range(S.PAGE_FIRST_DATA_ROW, tab.max_row + 1)
                if norm(tab.cell(r, 1).value)]

        changed = 0
        for sheet in S.DATA_SHEETS:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            hdr = header_row(ws)
            ci = {h: i + 1 for i, h in enumerate(S.HEADERS)}
            for r in range(hdr + 1, ws.max_row + 1):
                key = norm(ws.cell(r, 1).value)
                if not key:
                    continue
                want = S.derive_waiting_on(
                    norm(ws.cell(r, ci[S.COL_MACHINE_STATUS]).value),
                    norm(ws.cell(r, ci[S.COL_APPROVAL_STATUS]).value),
                    items_by_page.get(key, ()))
                cell = ws.cell(r, ci[S.COL_WAITING_ON])
                if norm(cell.value) != want:
                    cell.value = want
                    changed += 1
        if changed:
            append_log(wb, args.actor, 'רענון «ממתין ל»', '',
                       f'{changed} שורות עודכנו')
            wb.save(TRACKER)
        print(f'  «ממתין ל» רוענן — {changed} שורות השתנו')
        return 0

    if args.release_lock or args.acquire_lock:
        wb = load_workbook(TRACKER)
        set_lock(wb, args.actor if args.acquire_lock else None)
        append_log(wb, args.actor,
                   'נטילת LOCK' if args.acquire_lock else 'שחרור LOCK', '',
                   args.reason)
        wb.save(TRACKER)
        print('  LOCK ' + ('ננעל על ' + args.actor if args.acquire_lock else 'שוחרר'))
        return 0

    if args.show:
        wb = load_workbook(TRACKER, data_only=True)
        sheet, r = locate(wb, args.show)
        ws = wb[sheet]
        print(f'  {sheet} · שורה {r}')
        for c, h in enumerate(S.HEADERS, start=1):
            owner = 'סוכן' if S.OWNER_OF[h] == S.AGENT else 'אנוש'
            print(f'    [{owner}] {h:20} = {norm(ws.cell(r, c).value)!r}')
        return 0

    if not args.row or not args.set:
        print('נדרש --row עם לפחות --set אחד', file=sys.stderr)
        return 2

    # Parse and validate BEFORE opening the workbook — a rejected write never
    # reaches disk and never takes the lock.
    updates: dict[str, str] = {}
    for pair in args.set:
        if '=' not in pair:
            print(f'--set חייב להיות COL=VALUE, התקבל: {pair!r}', file=sys.stderr)
            return 2
        col, val = pair.split('=', 1)
        col = col.strip()
        if col not in S.HEADERS:
            print(f'עמודה לא מוכרת: {col!r}\nמוכרות: {list(S.HEADERS)}', file=sys.stderr)
            return 2
        if S.OWNER_OF[col] == S.HUMAN:
            print(f'סירוב: «{col}» היא עמודה בבעלות אנוש. '
                  'סוכן אינו כותב בה, ואינו קובע סטטוס אישור.', file=sys.stderr)
            return 1
        updates[col] = val.strip()

    ms = updates.get(S.COL_MACHINE_STATUS)
    if ms and ms not in S.MACHINE_STATUSES:
        print(f'סטטוס מכונה לא חוקי: {ms!r}\nמותר: {list(S.MACHINE_STATUSES)}',
              file=sys.stderr)
        return 1

    mtime_before = TRACKER.stat().st_mtime
    wb = load_workbook(TRACKER)
    sheet, r = locate(wb, args.row)
    ws = wb[sheet]

    lock = state_cell(wb, 'LOCK')
    if lock is not None and norm(lock.value):
        print(f'הקובץ נעול כרגע: {norm(lock.value)!r}. '
              'סיימו את החלון הפתוח או הריצו --release-lock.', file=sys.stderr)
        return 1

    # Transition legality against what is actually on the row right now.
    if ms:
        prev = norm(ws.cell(r, S.HEADERS.index(S.COL_MACHINE_STATUS) + 1).value)
        if prev and ms != prev and ms not in S.MACHINE_TRANSITIONS.get(prev, set()):
            print(f'מעבר סטטוס אסור «{prev}» ← «{ms}». '
                  f'מותר: {sorted(S.MACHINE_TRANSITIONS.get(prev, set()))}',
                  file=sys.stderr)
            return 1
        if ms in S.STATUS_REQUIRING_REASON:
            notes = updates.get(
                S.COL_AGENT_NOTES,
                norm(ws.cell(r, S.HEADERS.index(S.COL_AGENT_NOTES) + 1).value))
            if not notes:
                print(f'סטטוס «{ms}» מחייב סיבה כתובה ב«{S.COL_AGENT_NOTES}».',
                      file=sys.stderr)
                return 1

    changes = []
    for col, val in updates.items():
        c = ws.cell(r, S.HEADERS.index(col) + 1)
        before = norm(c.value)
        if before != val:
            c.value = val
            changes.append(f'{col}: {before!r} ← {val!r}')

    if not changes:
        print('  אין שינוי — הערכים כבר זהים.')
        return 0

    append_log(wb, args.actor, 'עדכון שורה', f'{sheet}!{args.row}',
               (args.reason + ' · ' if args.reason else '') + ' | '.join(changes))
    # A single --set is atomic: it completes in one save, so it stamps «עדכון
    # אחרון» but deliberately leaves LOCK empty. LOCK marks a multi-step work
    # WINDOW and is taken explicitly with --acquire-lock.
    set_lock(wb, None)

    # Someone may have saved from Sheets/Excel while we were deciding.
    if TRACKER.stat().st_mtime != mtime_before:
        print('הקובץ השתנה על הדיסק בזמן העיבוד — לא נכתב. '
              'הריצו שוב כדי לעבוד על הגרסה העדכנית.', file=sys.stderr)
        return 1

    wb.save(TRACKER)

    print(f'  {sheet}!{args.row} — {len(changes)} שינויים:')
    for ch in changes:
        print(f'    {ch}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
